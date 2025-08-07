from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

chrome_options = Options()
# chrome_options.add_argument("--headless")  # Run Chrome in headless mode
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36")
service = Service('C:/chromedriver-win32/chromedriver.exe')
driver = webdriver.Chrome(service=service,options=chrome_options)

import time
import re
time.sleep(5)

import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import openpyxl
from urllib.parse import urljoin
import spacy
import requests

class StaffData:

  def __init__(self, link_url='', name='', title='', twitter='', phone='', email='', dept=''):
    self.link_url = link_url
    self.name = name
    self.title = title
    self.twitter = twitter
    self.phone = phone
    self.email = email
    self.dept = dept
  
  def to_array(self):
    return [self.dept, self.name, self.link_url, self.title, self.phone, self.email, self.twitter]

class TrData:
  def __init__(self, tr_data=None):
    self.td_list = []
    if tr_data is not None:
      td_list = tr_data.select('td') or tr_data.select('th')
      # if len(td_list) > 1:
      #   th_data = tr_data.select_one('th')
      #   self.td_list.append(TdData(th_data))
      for td_data in td_list:
        self.td_list.append(TdData(td_data))
  
  def get_dept(self, orders):
    index = orders["department"]
    if index != -1 and index < len(self.td_list):
      return self.td_list[index].get_text()

    non_empty_depts = [td_data for td_data in self.td_list if td_data.get_text()]
    
    if len(non_empty_depts) <= 1:
      return get_title_from_category(non_empty_depts[0].td_data)
      # for td_data in non_empty_depts:
      #   text = td_data.td_data.get_text()
      #   doc = nlp(text)
      #   for ent in doc.ents:
      #     if ent.label_ == "PERSON":
      #       return ""
      # return get_title_from_category(non_empty_depts[0].td_data)
    
    return ""
  def is_empty(self):
    for td_data in self.td_list:
      if td_data.get_text():
        return False
    return True
  
  def is_header(self):
    patterns = ["Header", "Mail", "Name", "Title", "Phone", "Email", "Twitter"]
    match_len = 0
    for td_data in self.td_list:
      if td_data.get_text() in patterns:
        match_len += 1
    
    if match_len >= 3:
      return True
    
    return False
      
  def is_title(self, text):
    if text == '':
      return False
    patterns = ["Athletic", "Director", "Manager", "Sports", "Staff", "Defensive", "Offensive", "Receivers",
      "Safeties", "Broadcaster",
      "Coach", "Assistant", "Head", "Associate", "Compliance", "President",
      "Representative", "Senior", "Junior", "Registered", "Psychiatrist", "Primary",
      "Supervisor", "Coordinator", "Volunteer", "Student", "Athlete", "Trainer", "Physician",
      "Advisor", "Counselor", "Specialist", "Officer", "Recruiter", "Recruiting", "Recruitment"]
    
    sports_pattern = ["Acrobatics & Tumbling", "Archery", "Badminton", "Baseball", "Basketball",
    "Beach Volleyball", "Bowling", "Crew/Rowing", "Cycling/Biking", "Equestrian",
    "Fencing", "Field Hockey", "Football", "Golf", "Gymnastics", "Ice Hockey",
    "Lacrosse", "Martial Arts", "Polo", "Rifle/Shooting", "Rodeo", "Rugby",
    "Sailing", "Skiing", "Soccer", "Softball", "Sprint Football", "Squash",
    "Swimming & Diving", "Tennis", "Volleyball", "Water Polo", "Wrestling",
    "Swimming", "Diving", "Cheerleading", "Hockey", "Triathlon", "Cheerleader", "Dance",
    "Esports", "Handball", "Skating", "Waterski", "Athletic Staff", "Futsal",
    "Swim", "Dive", "Shotgun", "Stunt", "Surf", "Cheer & Stunt", "Cornhole",
    "Boxing", "Spirit Squad", "Cheer"
    "Climbing", "Club Sport", "Cross Country/Track & Field", "Fishing",
    "Flag Football", "JV Sport", "Logging Sport", "Marching Band/Band",
    "Outdoor Track & Field", "Indoor Track & Field", "Track and Field",
    "Motorsports", "Snowboarding"]
    
    patterns += sports_pattern
    return any(word in text for word in patterns)
  
  def is_phone(self, text):
    if text == '':
      return False
    if len(re.findall(r'\d', text)) > 4:
      return True
    return False
  
  def is_email(self, text):
    if text == '':
      return False
    email_pattern = re.compile(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$')
    return re.match(email_pattern, text)
  
  def is_name(self, text):
    if text == '':
      return False
    doc = nlp(text)
    for ent in doc.ents:
      if ent.label_ == "PERSON":
        return True
    
    if len(text.split()) >= 2 and not (self.is_email(text) or self.is_title(text) or self.is_phone(text)):
      return True
    
    return False

  def get_email_td(self, orders, specical_case = 0):
    index = orders["mail"] + specical_case
    if index != -1 and index < len(self.td_list):
      return self.td_list[index]
    for td_data in self.td_list:
      if self.is_email(td_data.get_text()):
        return td_data

    return TdData()

  def get_title_td(self, orders, specical_case = 0):
    index = orders["title"] + specical_case
    if index != -1 and index < len(self.td_list):
      return self.td_list[index]
    
    for td_data in self.td_list:
      text = td_data.get_text()
      if self.is_title(text):
        return td_data

    return TdData()
  
  def get_twitter_td(self, orders, special_case = 0):
    index = orders["twitter"] + special_case

    if index != -1 and index < len(self.td_list):
      return self.td_list[index]

    return TdData()
  
  def get_name_td(self, orders, special_case = 0):
    index = orders["name"] + special_case

    if index != -1 and index < len(self.td_list):
      return self.td_list[index]
    
    for td in self.td_list:
      text = td.get_text()
      if self.is_name(text):
        return td
    
    return TdData()
  
  def get_phone_td(self, orders, special_case = 0):
    index = orders["phone"] + special_case
    if index != -1 and index < len(self.td_list):
      return self.td_list[index]
    
    for td_data in self.td_list:
      text = td_data.get_text()
      if self.is_phone(text):
        return td_data

    return TdData()

class TdData:

  def __init__(self, td_data=None):
    self.td_data = td_data
  
  def get_text(self):
    if self.td_data is None:
      return ""
    text= self.td_data.text.strip() if self.td_data else ""
    text = re.sub(r'[\r\t\n]', '', text)
    return text
  
  def get_phone(self):
    if self.td_data is None:
      return ""
    for item in self.td_data.select("a"):
      if 'class' in item.attrs:
        if 'hide-on-large' in item['class']:
          return item.text.strip()
    
    return self.get_text()
  
  def get_link_url(self):
    if self.td_data is None:
      return ""
    link = self.td_data.select_one('a')
    link_url = ''
    if link is not None:
      link_url = urljoin(url, link.get('href'))
    return link_url
  
  def get_name_title(self):
    if self.get_text() is None:
      return "", ""
    name = self.td_data.select_one('a').text.strip()
    title = self.get_text().replace(name, "")
    title = re.sub(r'[\r\t\n]', '', title)
    return name, title

def get_title_from_category(element):
  if element is not None:
    for line in list(element.stripped_strings):
      t_list = re.sub(r'[\r\t\n]', '', line).split('|')
      for t in t_list:
        if t.strip() != '':
          return t.strip()
    t_list = re.sub(r'[\r\t\n]', '', element.text.strip()).split('|')
    for t in t_list:
      if t.strip() != '':
        return t.strip()
  return ''

def get_column_order_from_table(table):
  header_names = ["Name", "Title", "Position", "mail", "Phone", "Twitter", "Social", "Department"]
  td_list = []
  orders = {}
  print("I am in")

  try:
    head_element = table.select_one("thead")
    print(head_element)
    td_list = head_element.select_one("tr").select("th")
    orders["index"] = 0
  except:
    tr_list = table.select("tr")
    for index, tr in enumerate(tr_list):
      text = tr.text.strip()
      match_count = 0
      for header_name in header_names:
        if header_name.lower() in text.lower():
          match_count += 1
      if match_count >= 3:
        td_list = tr.select("td") or tr.select("th")
        orders["index"] = index
        break

  if td_list is None or len(td_list) == 0:
    return {
      "index": -1,
      "name": -1,
      "title": -1,
      "mail": -1,
      "phone": -1,
      "twitter": -1,
      "department": -1
    }

  text_list = [th.text.strip() for th in td_list]
  id_list = [th["id"] for th in td_list if "id" in th.attrs]
  
  orders_mapping = {
    "Name": "name",
    "Title": "title",
    "Position": "title",  # Combine "Title" and "Position"
    "mail": "mail",
    "Phone": "phone",
    "Twitter": "twitter",
    "Social": "twitter",  # Combine "Twitter" and "Social"
    "Department": "department"
  }

  for index, text in enumerate(text_list):
    for keyword, key in orders_mapping.items():
      if keyword.lower() in text.lower():
        orders[key] = index
  
  for index1, id in enumerate(id_list):
    for keyword, key in orders_mapping.items():
      if key in id and key not in orders:
        orders[key] = index1


  for key in orders_mapping.values():
    if key not in orders:
      orders[key] = -1
  return orders

def get_url_content(url):
  try:
    driver.get(url)
    wait = WebDriverWait(driver, 20)
    wait.until(EC.visibility_of_element_located((By.TAG_NAME, "table")))
    
    html_content = driver.page_source
  except:
    print("I am except")
    headers = {
      "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
    }
    page = requests.get(url, headers=headers)
    html_content = page.text
  soup = BeautifulSoup(html_content, "html.parser")
  return soup
  
def get_data_from_site_a(tables):
  member_list = []
  member_index = 0

  for index, table in enumerate(tables):
    if "role" in table.attrs and table["role"] == "presentation":
      continue
    try:
      if len(member_list) > 1:
        print("A_S_Type HD found membertable in index: ", index)
        break
      tr_list = table.select('tr')
      current_dept = ''

      orders = get_column_order_from_table(table)
      print(orders)

      for index1, tr in enumerate(tr_list):
        try:
          tr_data = TrData(tr)
          dept = tr_data.get_dept(orders)
          if dept != '':
            current_dept = dept
            print(current_dept)
            if orders["department"] == -1:
              continue
          if current_dept == '' or tr_data.is_empty():
            continue
          if index1 <= orders["index"] or tr_data.is_header():
            continue
          print(current_dept)
          staff_data = StaffData()
          staff_data.name = tr_data.get_name_td(orders).get_text()
          staff_data.link_url = tr_data.get_name_td(orders).get_link_url()
          staff_data.title = tr_data.get_title_td(orders).get_text()
          staff_data.phone = tr_data.get_phone_td(orders).get_phone()
          staff_data.email = tr_data.get_email_td(orders).get_text()
          staff_data.twitter = tr_data.get_twitter_td(orders).get_text() 
          staff_data.dept = current_dept
          member = [school_id, school_name] + staff_data.to_array()
          if member_index == 0:
            print(member)
          member_index += 1
          member_list.append(member)
        except Exception as e:
          print(str(e))
          continue
    except Exception as e:
      print(str(e))
      continue
  return member_list

def get_data_form_site_b(table_list, mode):
  member_list = []
  member_index = 0

  for table in table_list:
    if "role" in table.attrs and table["role"] == "presentation":
      continue
    try:
      dept = table.find_previous(lambda tag: tag.name and tag.name.startswith('h') or tag.name == 'center')
      while True:
        text = re.sub(r'[\r\t\n]', '', dept.text.strip())
        if text != '':
          break
        dept = dept.find_previous(lambda tag: tag.name and tag.name.startswith('h') or tag.name == 'center')
      dept = dept.text.strip()
      orders = get_column_order_from_table(table)
      print(orders)

      for index1, tr in enumerate(table.select('tr')):
        try:
          if index1 <= orders["index"]:
            continue
          staff_data = StaffData()
          tr_data = TrData(tr)

          if tr_data.is_empty():
            continue

          if mode == "BOTH_TITLE_NAME":
            staff_data.name, staff_data.title = tr_data.get_name_td(orders, 1).get_name_title()
          else:
            staff_data.name = tr_data.get_name_td(orders).get_text()
            staff_data.title = tr_data.get_title_td(orders).get_text()
          staff_data.link_url = tr_data.get_name_td(orders).get_link_url()
          staff_data.phone = tr_data.get_phone_td(orders).get_text()
          staff_data.email = tr_data.get_email_td(orders).get_text()
          staff_data.twitter = tr_data.get_twitter_td(orders).get_text()
          staff_data.dept = dept
          member = [school_id, school_name] + staff_data.to_array()
          if member_index == 0:
            print(member)
          member_index += 1
          member_list.append(member)
        except Exception as e:
          print(str(e))
          continue
    except Exception as e:
      print(str(e))
      continue
  return member_list

def get_data_from_site_c(url):
  api_url = urljoin(url, "api/v2/staff?$pageIndex=1&$pageSize=500&groupedByCategories=true&hash=2340353865518470")
  api_url = re.sub("/staff-directory", "", api_url)

  json_data = {}

  headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3",
    "Accept": "application/json, text/plain, */*",
    "Connection": "keep-alive",
    "Accept-Encoding": "gzip, deflate, br",
  }

  response = requests.get(api_url, headers=headers)
  
  if response.status_code == 200 or response.status_code == 202:
    json_data = response.json()
  else:
    print("Error: ", response.status_code)
    return []

  member_list = []
  member_index = 0

  for category in json_data["items"]:
    dept = category["groupName"]
    for member in category["groupItems"]:
      staff_data = StaffData()
      staff_data.name = member["firstName"] + " " + member["lastName"]
      staff_data.title = member["title"].strip()
      staff_data.link_url = urljoin(url, member["url"])
      staff_data.phone = member["phone"]
      staff_data.email = member["email"]
      staff_data.twitter = member["socialMedia"]["twitter"]
      staff_data.dept = dept
      member = [school_id, school_name] + staff_data.to_array()
      if member_index == 0:
        print(member)
      member_index += 1
      member_list.append(member)
  return member_list

def crawl_data(url, table_type):
  if table_type == 'type_b_c':
    return get_data_from_site_c(url)
  soup = get_url_content(url)
  tables = soup.select('table')
  member_list = []
  if table_type == 'type_a':
    print("I am in", len(tables))
    member_list = get_data_from_site_a(tables)
  elif table_type == 'type_b':
    member_list = get_data_form_site_b(tables, "NORMAL")
  elif table_type == 'type_b_b':
    member_list = get_data_form_site_b(tables, "BOTH_TITLE_NAME")
  return member_list

def drop_row_from_excel_file(url, school_ids):
  try:
    df = pd.read_excel(url)
    for index, row in df.iterrows():
      if row['school_id'] in school_ids:
        df.drop(index=index, inplace=True)
    df.to_excel(url, index=False)
  except Exception as e:
    print(str(e))
    return

final_columns = ['school_id', 'school_name', 'dept', 'name', 'person_profile_link', 'title', 'phone', 'email', 'twitter']
column_names = ['name', 'title', 'phone', 'email', 'twitter']
overwite_columns = ['school_id', 'school_name', 'is_enabled', 'staff_directory_url', 'is_structured', 'a_s_type', 'table_type', 'is_crawled']

input_excel_file = 'schools_10_11_to_crawl.xlsx'
sheet_name = 'Sheet1'

nlp = spacy.load("en_core_web_sm")

df = pd.read_excel(input_excel_file, sheet_name)

failed_school_id = []

try:
  drop_row_from_excel_file("staff_member_recrawl_10_12.xlsx", failed_school_id)
  wb_staff = openpyxl.load_workbook("staff_member_recrawl_10_12.xlsx")
  ws_staff = wb_staff.active
except Exception as e:
  print(str(e))
  wb_staff = openpyxl.Workbook()
  ws_staff = wb_staff.active
  ws_staff.append(final_columns)

new_data = []
for index, row in df.iterrows():
  try:
    school_id = int(row["school_id"])
    # if school_id <= 1500 and school_id not in failed_school_id:
    #   continue
    # if school_id != 470:
    #   continue
    url = row["staff_directory_url"]
    school_name = row["school_name"]

    table_type = row["table_type"]
    
    members = crawl_data(url, table_type)
    if len(members) > 0:
      print(f"==========================================={school_name} has {len(members)}, crawl success==========================================")
      for member in members:
        ws_staff.append(member)
      wb_staff.save(f'staff_member_recrawl_10_12.xlsx')
      row['is_crawled'] = True
      new_data.append(row.values.tolist())
    else:
      row['is_crawled'] = False
      new_data.append(row.values.tolist())
      print(f"==========================================={school_name} crawl faield ==========================================")
  except Exception as e:
    print(str(e))
    row['is_crawled'] = False
    new_data.append(row.values.tolist())
    if row["school_name"] is None:
      print(f'\n\n\n\nRow {index+1} {row["school_name"]} has no school ID\n\n\n')
    continue

driver.quit()

drop_row_from_excel_file("schools_1011_modified.xlsx", failed_school_id)

try:
  # Try to read the existing file
  new_df = pd.read_excel("schools_1011_modified.xlsx", sheet_name)
  # Append the new data and update new_df
  new_df = pd.concat([new_df, pd.DataFrame(new_data, columns=overwite_columns)], ignore_index=True)
  # Save the updated DataFrame to the Excel file
  new_df.to_excel("schools_1011_modified.xlsx", index=False)

except FileNotFoundError:
  # Handle the case where the file doesn't exist initially
  new_df = pd.DataFrame(new_data, columns=overwite_columns)
  new_df.to_excel("schools_1011_modified.xlsx", index=False)

except Exception as e:
  # Handle other exceptions
  print(f"An error occurred: {str(e)}")

